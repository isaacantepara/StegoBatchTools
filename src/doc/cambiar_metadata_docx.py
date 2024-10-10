from docx import Document
from datetime import datetime
import os
import time
import random
import openpyxl
import sys

def modificar_metadatos(archivo, archivo_excel):
    """Modifica los metadatos de un archivo DOCX."""
    doc = Document(archivo)
    print(f"Modificando metadatos para: {archivo}")

    # Leer el archivo Excel
    workbook = openpyxl.load_workbook(archivo_excel)
    worksheet = workbook.active

    # Elegir un autor aleatorio del archivo Excel
    row_index = random.randint(2, worksheet.max_row)  # Empieza en la fila 2 para evitar la cabecera
    autor_seleccionado = worksheet.cell(row=row_index, column=1).value
    fecha_captura = worksheet.cell(row=row_index, column=3).value

    print(f"Autor seleccionado: {autor_seleccionado}, Fecha captura: {fecha_captura}")

    line = autor_seleccionado.split(" ") 
    nombre_completo = f"{line[0]}_{line[1]}"

    # Convertir la cadena de texto a un objeto datetime
    if isinstance(fecha_captura, datetime):
        target_datetime = fecha_captura
    else:
        try:
            target_datetime = datetime.strptime(fecha_captura, "%d/%m/%Y")  # Formato de fecha en el Excel
        except ValueError as ve:
            print(f"Error en la conversión de fecha para {autor_seleccionado}: {ve}")
            return

    # Establecer el autor en el documento
    doc.core_properties.author = nombre_completo

    # Establecer metadatos con el objeto datetime
    doc.core_properties.created = target_datetime
    doc.core_properties.modified = target_datetime
    doc.core_properties.description = "lil" 

    # Guardar los cambios
    doc.save(archivo)

    # Actualizar las fechas del archivo con el timestamp
    timestamp = int(time.mktime(target_datetime.timetuple()))
    os.utime(archivo, (timestamp, timestamp))

if __name__ == "__main__":
    # Obtener la ruta de la carpeta de archivos DOCX y el archivo Excel desde los argumentos
    if len(sys.argv) != 3:
        print("Uso: python cambiar.py <ruta_carpeta_docx> <ruta_archivo_excel>")
        sys.exit(1)

    carpeta_docx = sys.argv[1]
    archivo_excel = sys.argv[2]

    print(f"Buscando archivos en: {carpeta_docx}")

    # Iterar sobre los archivos DOCX en la carpeta
    for archivo in os.listdir(carpeta_docx):
        if archivo.endswith(".docx"):
            ruta_archivo = os.path.join(carpeta_docx, archivo)
            try:
                modificar_metadatos(ruta_archivo, archivo_excel)
            except Exception as e:
                print(f"Error al modificar {archivo}: {e}")
