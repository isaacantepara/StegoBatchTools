import piexif
import openpyxl
import os
import sys
import random
from datetime import datetime

def modificar_metadatos_imagen(archivo, archivo_excel):
    """Modifica los metadatos de una imagen."""
    print(f"Modificando metadatos de: {archivo}")

    # Leer el archivo Excel
    workbook = openpyxl.load_workbook(archivo_excel)
    worksheet = workbook.active

    # Elegir un autor aleatorio del archivo Excel
    row_index = random.randint(2, worksheet.max_row)  # Empieza en la fila 2 para evitar la cabecera
    autor_seleccionado = worksheet.cell(row=row_index, column=1).value
    fecha_captura = worksheet.cell(row=row_index, column=3).value

    # Formatear el nombre del autor
    line = autor_seleccionado.split(" ")
    nombre_completo = f"{line[0]}_{line[1]}"

    # Convertir la fecha a un objeto datetime
    if isinstance(fecha_captura, datetime):
        target_datetime = fecha_captura
    else:
        target_datetime = datetime.strptime(fecha_captura, "%d/%m/%Y")  # Formato de fecha en el Excel

    # Cargar los datos EXIF
    try:
        exif_dict = piexif.load(archivo)
    except Exception as e:
        print(f"Error al cargar EXIF de {archivo}: {e}")
        return

    # Agregar o modificar el autor en los metadatos EXIF
    if '0th' not in exif_dict:
        exif_dict['0th'] = {}
    
    exif_dict['0th'][piexif.ImageIFD.Artist] = nombre_completo.encode('utf-8')

    # Establecer la fecha de captura
    fecha_str = target_datetime.strftime("%Y:%m:%d %H:%M:%S")
    exif_dict['0th'][piexif.ImageIFD.DateTime] = fecha_str.encode('utf-8')

    # Convertir de nuevo a bytes
    exif_bytes = piexif.dump(exif_dict)

    # Guardar la imagen con los nuevos metadatos
    try:
        with open(archivo, 'rb') as img_file:
            img_data = img_file.read()
        with open(archivo, 'wb') as img_file:
            img_file.write(img_data)
            img_file.write(exif_bytes)  # Aquí deberías modificar para agregar EXIF correctamente

    except Exception as e:
        print(f"Error al guardar {archivo}: {e}")

def main(ruta_imagenes, archivo_excel):
    """Función principal para modificar las imágenes en una carpeta."""
    for filename in os.listdir(ruta_imagenes):
        if filename.lower().endswith(('.jpg', '.jpeg')):
            ruta_archivo = os.path.join(ruta_imagenes, filename)
            modificar_metadatos_imagen(ruta_archivo, archivo_excel)

if __name__ == "__main__":
    if len(sys.argv) != 3:
        print("Uso: python test.py <ruta_imagenes> <archivo_excel>")
    else:
        ruta_imagenes = sys.argv[1]
        archivo_excel = sys.argv[2]
        main(ruta_imagenes, archivo_excel)
